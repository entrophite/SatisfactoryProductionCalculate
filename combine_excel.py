#!/usr/bin/env python3

import itertools
import os
from typing import Generator

import openpyxl


def formatted_string_generator(*fmts, **kw: list[tuple],
) -> Generator[tuple[str, str], None, None]:
	keys = list(kw.keys())
	for kv in itertools.product(*(kw[k] for k in keys)):
		yval = list()
		for i, fmt in enumerate(fmts):
			vals = dict(zip(keys, [v[i] for v in kv]))
			yval.append(fmt.format(**vals))
		yield tuple(yval)
	return


def tsv_to_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet,
	fname: str
) -> None:
	with open(fname, "r", encoding="utf-8") as fp:
		for r, line in enumerate(fp):
			for c, cell in enumerate(line.strip().split("\t")):
				if cell.startswith("="):
					cell = cell.replace("=", "-")
				try:
					cell = float(cell)
				except ValueError:
					pass
				worksheet.cell(row=r + 1, column=c + 1, value=cell)
	return


if __name__ == "__main__":
	wb = openpyxl.Workbook()
	for fname, ws_name in formatted_string_generator(
		"output/calc.max_power{waste}{conv}{sink}{sloop}.txt",
		"最大化发电-{waste}-{sink}-{conv}-{sloop}",
		waste=[
			(".waste_free", "无废料"),
			(".waste_prone", "有废料"),
		],
		sink=[
			(".sink_pluto", "钚回收"),
			(".ficsonium", "铀钚镄"),
		],
		conv=[
			("", "允许转化"),
			(".wo_conv", "无转化"),
		],
		sloop=[
			(".with_sloop", "有红石"),
			(".wo_sloop", "无红石"),
		],
	):
		if not os.path.exists(fname):
			continue
		ws = wb.create_sheet(title=ws_name)
		tsv_to_worksheet(ws, fname)

	for fname, ws_name in formatted_string_generator(
		"output/calc.max_power{waste}{sloop}.txt",
		"最大化发电-{waste}-{sloop}",
		waste=[
			(".waste_free", "无废料"),
			(".waste_prone", "有废料"),
		],
		sloop=[
			(".with_sloop", "有红石"),
			(".wo_sloop", "无红石"),
		],
	):
		if not os.path.exists(fname):
			continue
		ws = wb.create_sheet(title=ws_name)
		tsv_to_worksheet(ws, fname)

	for fname, ws_name in formatted_string_generator(
		"output/calc.max_point{oc}{sloop}.txt",
		"最大化点数-{oc}-{sloop}",
		oc=[
			(".oc_1", "降频1%"),
			(".oc_100", "基础100%"),
			(".oc_250", "超频250%"),
		],
		sloop=[
			(".with_sloop", "有红石"),
			(".wo_sloop", "无红石"),
		],
	):
		if not os.path.exists(fname):
			continue
		ws = wb.create_sheet(title=ws_name)
		tsv_to_worksheet(ws, fname)

	wb.remove(wb["Sheet"])
	wb.save("output/results.xlsx")
